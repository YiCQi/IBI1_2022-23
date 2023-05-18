# I am sorry for my poor code... It runs really slowly (1min).
# But believe me... It is valid.
from openpyxl import Workbook
from xml.dom.minidom import parse
import xml.dom.minidom

# Collect file
DOMTree = xml.dom.minidom.parse('go_obo.xml')
collection = DOMTree.documentElement
# Collect terms
terms = collection.getElementsByTagName('term')

# Use workbook to edit Excel
workbook = Workbook()
sheet = workbook.active
sheet['A1'] = 'id'
sheet['B1'] = 'name'
sheet['C1'] = 'definition'
sheet['D1'] = 'childnodes'
# Initial row index
row_index = 2

# Use recursion to count child nodes
def count_nodes(termid):
    nodes = 0
    for term_ in terms:
        is_a_ = term_.getElementsByTagName('is_a')
        for item in is_a_:
            if termid == item.firstChild.data:
                nodes += 1 + count_nodes(term_.getElementsByTagName('id')[0].firstChild.data)
    return nodes

# Get id, name, definition from term which has autophagosome in <defstr>
for term in terms:
    defstr = term.getElementsByTagName('defstr')[0].firstChild.data
    if 'autophagosome' in defstr:
        term_id = term.getElementsByTagName('id')[0].firstChild.data
        term_name = term.getElementsByTagName('name')[0].firstChild.data
        definition = defstr

        # Edit Excel sheet cell
        sheet.cell(row=row_index, column=1).value = term_id
        sheet.cell(row=row_index, column=2).value = term_name
        sheet.cell(row=row_index, column=3).value = definition
        sheet.cell(row=row_index, column=4).value = count_nodes(term_id)
        row_index += 1

# Save Excel file
workbook.save('autophagosome.xlsx')